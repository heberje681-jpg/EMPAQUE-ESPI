"""
Núcleo de conciliación de pallets ESPI.

Lee el archivo de LIQUIDACION (una pestaña por productor/lote, con varios
bloques diarios, cada uno con una tabla de pallets '#','CAL','cajas') y el
archivo RCF (una pestaña por manifiesto, con # PALLET, LOTE, CAJAS, CALIBRE,
PRODUCTOR, MANIFIESTO), cruza cada pallet listado en LIQUIDACION contra el
RCF por (LOTE, # PALLET), y escribe los resultados de vuelta en una copia
del archivo de LIQUIDACION: marca cada pallet (encontrado / con diferencia /
no encontrado) y agrega una tabla de conciliación consolidada por lote.

No modifica ninguna fórmula ni tabla existente del archivo original — solo
agrega marcas de color/comentarios sobre las celdas de pallets y una tabla
nueva, claramente rotulada, al final de cada hoja de productor.
"""
import re
from collections import defaultdict
from difflib import SequenceMatcher

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------

def _norm(s):
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().upper()


def _first_digits(s):
    """Extrae el primer número entero encontrado en un texto (para LOTE)."""
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return int(s)
    m = re.search(r"\d+", str(s))
    return int(m.group()) if m else None


def _name_similarity(a, b):
    return SequenceMatcher(None, _norm(a), _norm(b)).ratio()


PRODUCTOR_MATCH_THRESHOLD = 0.45


def _same_productor(rec_productor, sheet_productor):
    """LOTE no es único por productor (varios productores pueden compartir un
    mismo número de lote físico), así que siempre se valida también el nombre."""
    if not sheet_productor:
        return True
    return _name_similarity(rec_productor, sheet_productor) >= PRODUCTOR_MATCH_THRESHOLD


# ---------------------------------------------------------------------------
# RCF: extracción de todos los manifiestos
# ---------------------------------------------------------------------------

_RCF_KEYS = {
    "FECHA": "FECHA", "PALLET": "PALLET", "ETIQUETA": "ETIQUETA", "LOTE": "LOTE",
    "CAJA": "CAJAS", "CALIBRE": "CALIBRE", "PRODUCTOR": "PRODUCTOR",
    "COMENTARIO": "COMENTARIOS", "MANIFIESTO": "MANIFIESTO", "VARIEDAD": "VARIEDAD",
}


def _find_rcf_header(ws):
    for row in ws.iter_rows(min_row=1, max_row=3):
        found = {}
        for c in row:
            nv = _norm(c.value)
            if not nv:
                continue
            for key, canon in _RCF_KEYS.items():
                if key in nv and canon not in found:
                    found[canon] = c.column
        if "PALLET" in found and "LOTE" in found and "CAJAS" in found:
            return row[0].row, found
    return None, None


def extract_rcf_records(wb, excluded_sheets=()):
    """Devuelve lista de dicts: sheet, row, pallet, lote, cajas, calibre, productor, manifiesto.

    Algunas pestañas de este archivo (típicamente una llamada "RCF") son un
    resumen maestro que repite filas que ya aparecen en las pestañas
    individuales de cada manifiesto — pero también puede contener manifiestos
    que todavía no tienen su propia pestaña. Por eso no se excluyen pestañas:
    se leen todas y se deduplican registros idénticos al final.
    """
    excluded_norm = {_norm(s) for s in excluded_sheets}
    records = []
    skipped_sheets = []
    for sn in wb.sheetnames:
        if _norm(sn) in excluded_norm:
            continue
        ws = wb[sn]
        hr, headers = _find_rcf_header(ws)
        if hr is None:
            skipped_sheets.append(sn)
            continue
        for r in range(hr + 1, ws.max_row + 1):
            pallet = ws.cell(row=r, column=headers["PALLET"]).value
            if pallet is None:
                continue
            lote = ws.cell(row=r, column=headers["LOTE"]).value
            cajas = ws.cell(row=r, column=headers["CAJAS"]).value
            calibre = ws.cell(row=r, column=headers.get("CALIBRE")).value if "CALIBRE" in headers else None
            productor = ws.cell(row=r, column=headers.get("PRODUCTOR")).value if "PRODUCTOR" in headers else None
            manifiesto = ws.cell(row=r, column=headers.get("MANIFIESTO")).value if "MANIFIESTO" in headers else sn
            try:
                pallet = int(pallet)
            except (TypeError, ValueError):
                continue
            lote_n = _first_digits(lote)
            records.append(dict(
                sheet=sn, row=r, pallet=pallet, lote=lote_n, lote_raw=lote,
                cajas=cajas, calibre=calibre, productor=productor,
                manifiesto=manifiesto if manifiesto is not None else sn,
            ))

    # Deduplicar registros idénticos que vienen de una pestaña resumen y de
    # su pestaña individual correspondiente, priorizando el registro que NO
    # viene de una pestaña llamada "RCF" a secas (preferimos la fuente
    # específica del manifiesto cuando existe).
    def dedup_key(rec):
        return (rec["pallet"], rec["lote"], rec["cajas"], _norm(rec["calibre"]),
                _norm(rec["productor"]), rec["manifiesto"])

    best = {}
    for rec in records:
        k = dedup_key(rec)
        if k not in best:
            best[k] = rec
        elif _norm(best[k]["sheet"]) == "RCF" and _norm(rec["sheet"]) != "RCF":
            best[k] = rec  # preferir la pestaña específica del manifiesto

    return list(best.values()), skipped_sheets


def index_rcf(records):
    idx = defaultdict(list)
    for rec in records:
        idx[(rec["lote"], rec["pallet"])].append(rec)
    return idx


# ---------------------------------------------------------------------------
# LIQUIDACION: detección de bloques, lote/productor y tablas de pallets
# ---------------------------------------------------------------------------

def find_pallet_tables(ws):
    """Ubica cada grupo de columnas (#, CAL, cajas) en toda la hoja."""
    tables = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "#":
                r, c = cell.row, cell.column
                nxt = ws.cell(row=r, column=c + 1).value
                if isinstance(nxt, str) and nxt.strip().upper().startswith("CAL"):
                    tables.append((r, c))
    return tables


def extract_pallet_rows(ws, header_row, col):
    """Lee pallets debajo de un header (#, CAL, cajas) hasta 2 filas vacías seguidas."""
    pallets = []
    r = header_row + 1
    blanks = 0
    while r - header_row < 20:
        v = ws.cell(row=r, column=col).value
        if v is None:
            blanks += 1
            if blanks > 1:
                break
            r += 1
            continue
        blanks = 0
        if isinstance(v, (int, float)):
            cal = ws.cell(row=r, column=col + 1).value
            caj = ws.cell(row=r, column=col + 2).value
            pallets.append(dict(pallet=int(v), calibre=cal, cajas=caj,
                                 row=r, col=col, coord=ws.cell(row=r, column=col).coordinate))
        r += 1
    return pallets


def find_sheet_lote_and_productor(ws, sheet_name):
    """Busca la(s) celda(s) 'LOTE:' y 'PROPIETARIO DE REZAGA' / VARIEDAD en la hoja."""
    lote_vals, productor_vals = [], []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                uv = cell.value.upper()
                if "LOTE" in uv and "LOTE" == uv.strip().rstrip(":").strip():
                    for c in range(cell.column + 1, cell.column + 6):
                        v = ws.cell(row=cell.row, column=c).value
                        if v is not None:
                            lote_vals.append(v)
                            break
                if "PROPIETARIO" in uv:
                    for c in range(cell.column + 1, cell.column + 6):
                        v = ws.cell(row=cell.row, column=c).value
                        if v is not None:
                            productor_vals.append(v)
                            break

    lote_number = None
    for v in lote_vals:
        n = _first_digits(v)
        if n is not None:
            lote_number = n
            break
    if lote_number is None:
        # usar el nombre de la pestaña como respaldo (ej. "8.1" -> lote 8)
        lote_number = _first_digits(sheet_name)

    productor = productor_vals[0] if productor_vals else sheet_name
    return lote_number, productor


# ---------------------------------------------------------------------------
# Conciliación por hoja de productor
# ---------------------------------------------------------------------------

def reconcile_sheet(ws, sheet_name, rcf_index, rcf_records_by_lote):
    lote, productor = find_sheet_lote_and_productor(ws, sheet_name)
    tables = find_pallet_tables(ws)

    all_pallets = []
    for r, c in tables:
        all_pallets.extend(extract_pallet_rows(ws, r, c))

    seen_coords = set()
    unique_pallets = []
    for p in all_pallets:
        if p["coord"] in seen_coords:
            continue
        seen_coords.add(p["coord"])
        unique_pallets.append(p)

    matched, mismatched, not_found = [], [], []
    listed_pallet_numbers = set()

    for p in unique_pallets:
        listed_pallet_numbers.add(p["pallet"])
        candidates = rcf_index.get((lote, p["pallet"]), [])
        # Un mismo número de lote puede tener varios productores (ej. lote
        # compartido); si hay varios candidatos, se prioriza el que coincide
        # con el nombre del productor de esta hoja.
        candidates = sorted(candidates, key=lambda rc: -_name_similarity(rc["productor"], productor))
        if not candidates:
            not_found.append(p)
            continue
        rec = candidates[0]
        p["rcf_match"] = rec
        # Solo se compara cuando el dato local está capturado; muchas hojas
        # no registran cajas por pallet (se asume la caja estándar del lote).
        same_calibre = (p["calibre"] is None) or (_norm(rec["calibre"]) == _norm(p["calibre"]))
        same_cajas = (p["cajas"] is None) or (rec["cajas"] == p["cajas"])
        if same_calibre and same_cajas:
            matched.append(p)
        else:
            mismatched.append(p)

    surplus = []
    if lote is not None:
        for rec in rcf_records_by_lote.get(lote, []):
            if rec["pallet"] in listed_pallet_numbers:
                continue
            if not _same_productor(rec["productor"], productor):
                continue  # pertenece a otro productor que comparte el mismo # de lote
            surplus.append(rec)

    # resumen por calibre
    by_calibre = defaultdict(lambda: dict(listados=0, cajas_listadas=0, encontrados=0,
                                           cajas_confirmadas=0, con_diferencia=0,
                                           no_encontrados=0, faltantes_pallets=[]))
    for p in unique_pallets:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        row = by_calibre[cal]
        row["listados"] += 1
        row["cajas_listadas"] += p["cajas"] or 0
    for p in matched:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["encontrados"] += 1
        by_calibre[cal]["cajas_confirmadas"] += p["rcf_match"]["cajas"] or 0
    for p in mismatched:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["con_diferencia"] += 1
        by_calibre[cal]["cajas_confirmadas"] += p["rcf_match"]["cajas"] or 0
    for p in not_found:
        cal = _norm(p["calibre"]) or "SIN CALIBRE"
        by_calibre[cal]["no_encontrados"] += 1
        by_calibre[cal]["faltantes_pallets"].append(p["pallet"])

    return dict(
        sheet=sheet_name, lote=lote, productor=productor,
        total_pallets_listados=len(unique_pallets),
        matched=matched, mismatched=mismatched, not_found=not_found, surplus=surplus,
        by_calibre=dict(by_calibre),
        pallet_table_locations=tables,
    )


# ---------------------------------------------------------------------------
# Escritura de resultados en la hoja (marcas + tabla de conciliación)
# ---------------------------------------------------------------------------

FILL_OK = PatternFill("solid", fgColor="C6EFCE")
FILL_WARN = PatternFill("solid", fgColor="FFEB9C")
FILL_BAD = PatternFill("solid", fgColor="FFC7CE")
FONT_OK = Font(color="006100")
FONT_WARN = Font(color="9C6500")
FONT_BAD = Font(color="9C0006")

THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _mark_pallet_cells(ws, result):
    for p in result["matched"]:
        cell = ws.cell(row=p["row"], column=p["col"])
        cell.fill = FILL_OK
        cell.font = FONT_OK
        cell.comment = Comment(f"OK - encontrado en manifiesto {p['rcf_match']['manifiesto']} "
                                f"({p['rcf_match']['sheet']})", "Conciliación ESPI")
    for p in result["mismatched"]:
        cell = ws.cell(row=p["row"], column=p["col"])
        cell.fill = FILL_WARN
        cell.font = FONT_WARN
        rec = p["rcf_match"]
        cell.comment = Comment(
            f"DIFERENCIA - en manifiesto {rec['manifiesto']} ({rec['sheet']}) "
            f"figura calibre {rec['calibre']} / {rec['cajas']} cajas",
            "Conciliación ESPI")
    for p in result["not_found"]:
        cell = ws.cell(row=p["row"], column=p["col"])
        cell.fill = FILL_BAD
        cell.font = FONT_BAD
        cell.comment = Comment("NO ENCONTRADO en ningún manifiesto RCF para este lote", "Conciliación ESPI")


def _write_summary_table(ws, result):
    # ubicar la fila libre después de todo el contenido existente
    start_row = ws.max_row + 3
    title_row = start_row

    ws.cell(row=title_row, column=1, value=f"CONCILIACIÓN AUTOMÁTICA — LOTE {result['lote']} "
                                            f"({result['productor']})").font = Font(bold=True, size=12)
    r = title_row + 1
    ws.cell(row=r, column=1, value=f"Total pallets listados en hoja: {result['total_pallets_listados']}")
    r += 1
    ws.cell(row=r, column=1,
            value=f"Encontrados OK: {len(result['matched'])}  |  "
                  f"Con diferencia: {len(result['mismatched'])}  |  "
                  f"No encontrados: {len(result['not_found'])}  |  "
                  f"Sobrantes en RCF (no listados aquí): {len(result['surplus'])}")
    r += 2

    headers = ["CALIBRE", "PALLETS LISTADOS", "CAJAS LISTADAS", "PALLETS ENCONTRADOS",
               "CAJAS CONFIRMADAS (RCF)", "CON DIFERENCIA", "NO ENCONTRADOS", "# PALLETS FALTANTES"]
    header_row = r
    for i, h in enumerate(headers):
        c = ws.cell(row=header_row, column=1 + i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    r += 1

    tot_listados = tot_cajas_list = tot_enc = tot_caj_conf = tot_dif = tot_no_enc = 0
    for cal in sorted(result["by_calibre"].keys()):
        row = result["by_calibre"][cal]
        vals = [cal, row["listados"], row["cajas_listadas"], row["encontrados"],
                row["cajas_confirmadas"], row["con_diferencia"], row["no_encontrados"],
                ", ".join(str(x) for x in row["faltantes_pallets"]) if row["faltantes_pallets"] else "—"]
        for i, v in enumerate(vals):
            c = ws.cell(row=r, column=1 + i, value=v)
            c.border = BORDER
            if i == 6 and row["no_encontrados"] > 0:
                c.fill = FILL_BAD
        tot_listados += row["listados"]; tot_cajas_list += row["cajas_listadas"]
        tot_enc += row["encontrados"]; tot_caj_conf += row["cajas_confirmadas"]
        tot_dif += row["con_diferencia"]; tot_no_enc += row["no_encontrados"]
        r += 1

    tot_vals = ["TOTAL", tot_listados, tot_cajas_list, tot_enc, tot_caj_conf, tot_dif, tot_no_enc, ""]
    for i, v in enumerate(tot_vals):
        c = ws.cell(row=r, column=1 + i, value=v)
        c.font = Font(bold=True)
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor="D9E1F2")
    r += 2

    if result["surplus"]:
        ws.cell(row=r, column=1, value="Pallets encontrados en RCF para este lote pero NO listados en esta hoja "
                                        "(posible sobrante / pallet no registrado en rezaga):").font = Font(bold=True)
        r += 1
        headers2 = ["# PALLET", "CALIBRE", "CAJAS", "PRODUCTOR (RCF)", "MANIFIESTO"]
        for i, h in enumerate(headers2):
            c = ws.cell(row=r, column=1 + i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="ED7D31")
            c.border = BORDER
        r += 1
        for rec in sorted(result["surplus"], key=lambda x: x["pallet"]):
            vals = [rec["pallet"], rec["calibre"], rec["cajas"], rec["productor"], rec["manifiesto"]]
            for i, v in enumerate(vals):
                ws.cell(row=r, column=1 + i, value=v).border = BORDER
            r += 1

    for col_idx in range(1, 9):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 20


def apply_results_to_workbook(wb, results_by_sheet):
    for sheet_name, result in results_by_sheet.items():
        ws = wb[sheet_name]
        _mark_pallet_cells(ws, result)
        _write_summary_table(ws, result)


# ---------------------------------------------------------------------------
# Punto de entrada de alto nivel
# ---------------------------------------------------------------------------

def run_reconciliation(liquidacion_path, rcf_path, output_path,
                        excluded_sheets=("RCF",)):
    wb_rcf_data = openpyxl.load_workbook(rcf_path, data_only=True)
    rcf_records, skipped_rcf_sheets = extract_rcf_records(wb_rcf_data, excluded_sheets=excluded_sheets)
    rcf_index = index_rcf(rcf_records)
    rcf_by_lote = defaultdict(list)
    for rec in rcf_records:
        rcf_by_lote[rec["lote"]].append(rec)

    wb_liq_data = openpyxl.load_workbook(liquidacion_path, data_only=True)
    wb_liq_out = openpyxl.load_workbook(liquidacion_path, data_only=False)

    results = {}
    for sn in wb_liq_data.sheetnames:
        ws = wb_liq_data[sn]
        tables = find_pallet_tables(ws)
        if not tables:
            continue  # hoja sin tabla de pallets reconocible, se omite
        result = reconcile_sheet(ws, sn, rcf_index, rcf_by_lote)
        results[sn] = result

    apply_results_to_workbook(wb_liq_out, results)
    wb_liq_out.save(output_path)

    return results, skipped_rcf_sheets
